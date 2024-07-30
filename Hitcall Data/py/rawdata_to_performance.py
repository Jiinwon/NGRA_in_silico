import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def create_assay_excel(input_file_path, output_file_path, template_file_path):
    # Load the input Excel file
    input_file = pd.ExcelFile(input_file_path)
    df = input_file.parse(input_file.sheet_names[0])

    # Load the template Excel file to match styles
    template_workbook = openpyxl.load_workbook(template_file_path)
    template_sheet = template_workbook.active

    # Define the required styles
    default_font = Font(name='맑은 고딕', size=11)
    bold_font = Font(name='맑은 고딕', size=11, bold=True)
    default_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Function to apply default styles including borders
    def apply_default_style_with_border(cell, bold=False):
        cell.font = bold_font if bold else default_font
        cell.alignment = default_alignment
        cell.border = border_style

    # Create a new Excel file with the specified formatting
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        for assay in df['Assays'].unique():
            # Filter the relevant rows for the current assay and sort by AUC
            df_assay_specific = df[df['Assays'] == assay][
                ['Algorithm', 'FP', 'AUC', 'F-Measure', 'Accuracy', 'Precision', 'Recall']].sort_values(by='AUC',
                                                                                                        ascending=False)

            # Write the sorted DataFrame to a new sheet in the output Excel file
            df_assay_specific.to_excel(writer, sheet_name=assay, index=False)

            # Load the sheet to apply styles
            workbook = writer.book
            worksheet = writer.sheets[assay]

            # Apply styles to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    apply_default_style_with_border(cell, bold=(cell.row == 1))

            # Set column widths
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 10

            # Apply yellow highlight to the row with the highest AUC value
            max_auc_index = df_assay_specific['AUC'].idxmax()
            max_row = df_assay_specific.index.get_loc(
                max_auc_index) + 2  # +2 to account for header and 1-based indexing in Excel

            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for col in range(1, len(df_assay_specific.columns) + 1):
                worksheet.cell(row=max_row, column=col).fill = fill


# 현재 파일(1.py)의 디렉토리 경로를 가져옴
current_dir = os.path.dirname(__file__)

# folder1에 있는 data.txt 파일의 경로를 설정
input_file_path = os.path.join(current_dir, '..', 'raw performance data', '6.xlsx')
output_file_path = os.path.join(current_dir, '..', 'Performance', 'updated_output_01.xlsx')
template_file_path = os.path.join(current_dir, '..', 'Performance', '퍼포먼스 저장9.xlsx')


create_assay_excel(input_file_path, output_file_path, template_file_path)
